import openpyxl
from openpyxl import workbook
import json
import pandas as pd
from FUNCIONES import Volver_Menu

def TXTaLista(opcion):
        diccionarios = []
        with open('info.txt', 'r') as file:
            # Lee cada línea del archivo
            for linea in file:
                try:
                    # Convierte la línea en un diccionario usando json.loads
                    diccionario_actual = json.loads(linea.strip())
                    diccionarios.append(diccionario_actual)
                except json.JSONDecodeError as e:
                    print(f"Error al decodificar JSON en la línea: {linea}")
                    print(e)
        if opcion == 1:    
            Crear_Excel(diccionarios)
        elif opcion ==2:
            return(diccionarios)

def Crear_Excel(datos):
    # Crea un nuevo libro de trabajo y selecciona la hoja activa
    libro = openpyxl.Workbook()
    hoja = libro.active
    # Encabezados
    encabezados = list(datos[0].keys())
    for col, encabezado in enumerate(encabezados, 1):
        hoja.cell(row=1, column=col, value=encabezado)
    # Datos
    for fila, persona in enumerate(datos, 2):
        for col, valor in enumerate(persona.values(), 1):
            hoja.cell(row=fila, column=col, value=valor)
    # Guarda el libro de trabajo en el archivo
    libro.save("info.xlsx")
    print("'info.xlsx' se ha actualizado")
    a = int(input("¿Desea abrir el archivo excel?\n1.Si\n2.No\n-"))
    while a < 1 or a > 2:
        print("ERROR: opcion no valida")
        a = int(input("¿Desea abrir el archivo excel?\n1.Si\n2.No\n-"))
    if a == 1:
        nombre_archivo = 'info.xlsx'
        # Leer el archivo Excel
        datos_excel = pd.read_excel(nombre_archivo)
        # Mostrar los datos
        print(datos_excel)
        Volver_Menu()
    else:
        Volver_Menu()
